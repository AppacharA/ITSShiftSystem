import MySQLdb
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import daySchedulingFunctions
import AuxiliaryFunctions
import re

def adminAddWorker(): 
	name = input("Type the first and last name of the worker you want to add here, followed by the username and the class year, separated by commas. For example, 'John Doe, jdoe, 2018'")
	inputArray = name.split(,)
	print("Here's what I have. Name: " + inputArray[0] + ", Username: " + inputArray[1] + ", Class year: " + inputArray[2])
	if input("Is that correct? (y/n)") == "y":
		nameArray = inputArray[0].split() #get firstname and last name
		workerTuple = (nameArray[0], nameArray[1], inputArray[1], inputArray[2])
		print (addWorker(workerTuple))

	else:
		print ("It's all right, we all make mistakes. Try again.")

def adminRemoveWorker():
	name = input("Type the first and last name of the worker you wish to remove.")
	nameArray = name.split()
	print ("I have: first name '" + nameArray[0] +"' and last name '"+ nameArray[1] + "'.")
	if input("Is that correct? (y/n)") == "y":
		workerTuple = (nameArray[0], nameArray[1])
		removeWorker(workerTuple)



def addWorker(workerTuple): #take in information of worker as a tuple (FirstName, LastName, username, classyear)
	cursor = db.cursor() #create a new cursor
	
	#check if employee table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
.		"WHERE table_schema = 'test' "
		"AND table_name = 'employeeinfo'"
		"LIMIT 1")


	if (cursor.rowcount == 0): #If your query returned a row, then the table exists
		print ("There is no table of Employees! Consider creating one from scratch.")

	else:

		query = ("INSERT INTO employeeinfo "
			"(firstName, lastName, username, classYear) VALUES (%s, %s, %s, %s)")

		#data = (workerName[0], workerName[1])
		cursor.execute(query, workerTuple)
		
		db.commit()
		return "Worker successfully added." #MAYBE INPUT BETTER TEST OF VALIDITY?

	cursor.close()

def removeWorker(workernameTuple): #workername must be a tuple
	cursor = db.cursor() #create a new cursor
	
	#check if employee table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = 'test' "
		"AND table_name = 'employeeinfo'"
		"LIMIT 1")

	if (cursor.rowcount != 0): #query returns a row, table must exist
		query = ("DELETE FROM employeeinfo "
					"WHERE firstName = %s " 
					"AND lastName = %s")

		cursor.execute(query, workernameTuple)

		db.commit()
	cursor.close()

def initializeEmployeeInfoTable(database, workerlist): #takes in all the workers and makes a table out of them
	#initialize cursor
	cursor = database.cursor()
	
	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = 'test' "
		"AND table_name = 'employeeinfo'"
		"LIMIT 1")


	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table of employees!")

	else: #if your query did not return a row, then the table does not exist.
		#first, create the table.
		#
		cmd = ("CREATE TABLE employeeinfo "
				"(id SMALLINT NOT NULL AUTO_INCREMENT, "
				"firstName VARCHAR(50), "	#Table parameter one: Names must be within 50 characters in length
				"lastName VARCHAR(50), "
				"username VARCHAR(25), "	#Check with Dave on maximum size a username can be to be efficient...also doublecheck if varchar is the best type for a name.		
				"phoneNumber VARCHAR(15), "
				"classYear YEAR(4), "
				"PRIMARY KEY (id))")

		cursor.execute(cmd)
		cursor.execute("ALTER TABLE employeeInfo AUTO_INCREMENT = 1001")
		#now we populate the employee table
		insertCmd = ("INSERT INTO employeeinfo (firstName, lastName)"
					"VALUES (%s, %s)"
			)

		for key in workerList: #remember, workerList is a dictionary of workers
			
			workerTuple = AuxiliaryFunctions.getStudent(workerList, key)
			
			cursor.execute(insertCmd, workerTuple)

		cursor.close()
		database.commit()







def initializeShiftList(database): 
	#pass in startDate and endDate as Python dateTime objects
	dateholder = input("When does the term start? (mm/dd/yyyy): ")
	startDate = datetime.strptime(dateholder, "%m/%d/%Y")
	# dateHolder = input("When does the term end? (mm/dd/yyyy): ")
	# endDate = datetime.strptime(dateholder, "%m/%d/%Y")
	#initialize cursor
	cursor = database.cursor()
	print ("hello")
	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = 'test' "
		"AND table_name = 'ShiftList'"
		"LIMIT 1")

	print(cursor._last_executed)
	print (cursor.rowcount)
	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table of shifts!")
	
	else: #Now we create the table if none exists
		cmd = ("CREATE TABLE ShiftList "
			"(id SMALLINT NOT NULL AUTO_INCREMENT PRIMARY KEY, "
			"checkinTime TIME,"
			"Location VARCHAR(25) NOT NULL DEFAULT 'CMC', "
			"Date DATE NOT NULL, " #A DATE must be input in the form YYYY/MM/DD
			"Day VARCHAR(15), "
			"subRequested BOOLEAN NOT NULL DEFAULT '0' )")
		

		cursor.execute(cmd)
		cursor.close()




		

		currentDate = startDate
		while (currentDate < datetime(2018, 4, 2)):
			if currentDate.strftime("%A") == "Sunday" or currentDate.strftime("%A") == "Saturday":
				daySchedulingFunctions.buildWkndSchedule(database, currentDate)
				print ("GROUNDBEEF")
			elif currentDate.strftime("%A") == "Monday" or currentDate.strftime("%A") == "Wednesday":
				print ("HELLO")
				daySchedulingFunctions.buildMonWedSchedule(database, currentDate)
			elif currentDate.strftime("%A") == "Tuesday" or currentDate.strftime("%A") == "Thursday":
				print ("BEEP")
				daySchedulingFunctions.buildTueThuSchedule(database, currentDate)

			elif currentDate.strftime("%A") == "Friday":
				print ("MEEP")
				daySchedulingFunctions.buildFriSchedule(database, currentDate)

			currentDate = currentDate + timedelta(days = 1)
			
		#commit all changes to database
		database.commit()
			#print (currentDate)
		#having created the table, we populate the table. (NOTE: For ease of readability, I've shifted the methods that create the shifts for each individual day off to another file)
		
		# while currentDate != endDate:
		# 	if currentDate

		

		#Get the start date of the term:


def initializeShiftEmployeeLinker(database, listofWorkers):
	#we're assuming that shiftdb and employeedb have already been created.

	#initialize cursor
	cursor = database.cursor()
	
	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = 'test' "
		"AND table_name = 'shiftEmployeeLinker'"
		"LIMIT 1")


	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table linking employees and shifts")

	else:
		#shift linker table does not exist, we create it first.
	
		cursor.execute("CREATE TABLE shiftEmployeeLinker "
			"(employeeID SMALLINT NOT NULL, "
			"shiftID SMALLINT NOT NULL, "
			"subRequested BOOL NOT NULL DEFAULT 0, "
			"checkedIn BOOL NOT NULL DEFAULT 0, "
			"checkinTime TIME,"
			"FOREIGN KEY (employeeID) REFERENCES employeeinfo(id) ON DELETE CASCADE, "
			"FOREIGN KEY (shiftID) REFERENCES shiftList(id) ON DELETE CASCADE, "
			"PRIMARY KEY (employeeID, shiftID))"

			)


		workerName = ""
		#now to populate the shiftlinker table. We will cycle through the entire Schedule Spreadsheet and match each worker to their shift.
		#filename = input("Place ITS Schedule Spreadsheet in the same folder as this program, then type its filename here (e.g Schedule.xlsx): ")
		filename = "Schedule.xlsx"
		wb = load_workbook(filename)
		

		timeRegex = re.compile('[^a-zA-z.\s]') #This regex matches any string that contains a number i.e a non-name cell.
		for col in range(1,15): #this corresponds to columns A through N
			for row in range(2,52): #this corresponds to the maximum row that the schedule spreadsheet currently has. Modify as necessary.
				cellValue = scheduleSheet.cell(row = row, column = col).value
				#print (str(row), str(col))
				#print (cellValue)
				if cellValue != None:
					cellValue = str(cellValue)	
					if timeRegex.search(cellValue) == None: #check that you're not actually dealing with a blank cell or a Time Cell, using the regex initialized above.
						Day = AuxiliaryFunctions.getDay(col) #Get the day of the shift as a string.
						location = AuxiliaryFunctions.getLocation(col) #get the location of the shift.
						shiftTime = AuxiliaryFunctions.getShiftTime(row, col)
						#now to get the name of the worker (and thus their identity from the employeeTable)
						#print (cellName)
						workerName = AuxiliaryFunctions.getStudent(listofWorkers, cellValue)
						print (workerName)

						
						employeeQuery = ("SELECT id " 	#TODO: FIGURE OUT AN ELEGANT WAY TO SOPHIA PROCESS
								"FROM employeeInfo "
								"WHERE firstName = %s "
								"AND lastName = %s"
								)

						
						
						#workerName is a tuple of the form (firstName, lastName)
						
						cursor.execute(employeeQuery, workerName)
						result = cursor.fetchone()
						#for some reason cursor returns results as a tuple so....
						print (result)
						employeeid = result[0]

						print ("Employee: " + str(employeeid))
						
						
						#If we have a worker, we must also have a shift!
						shiftQuery = ("SELECT id " 
								"FROM ShiftList "
								"WHERE Day = %s "
								"AND checkinTime = %s "
								"AND Location = %s"
								)
						cursor.execute(shiftQuery, (Day, shiftTime, location))
						result = cursor.fetchone()
						shiftid = result[0]
						
						#having gotten an employeeID and a shiftID, we append things to the shiftlinker table.
						inputQuery = ("INSERT INTO shiftEmployeeLinker (employeeID, shiftID)"
										"VALUES (%s, %s)")

						values = (employeeid, shiftid)

						cursor.execute(inputQuery, values)

	cursor.close()

	database.commit()






def initializeDB(database, termDBname): #here, database is the overall database of schedules, and termdbname will be the termly schedule.
	cursor = database.cursor()
	#name = (termDBname,) #a one-tuple of names!
	#name = input("What do you want to name this Table Schedule?")
	now = datetime.now() 
	if now.month > 2 and now.month < 7:
		name = "Spring" + str(now.year)
	elif now.month >= 9 and now.month < 11:
		name = "Fall" + str(now.year)
	elif now.month >= 11 and now.month <= 2:
		name = "Winter" + str(now.year)
	else:
		print ("You don't even know who's working for you next year! Nice Try.")

	cmd = ("create database if not exists " + name) #YOU CANNOT PARAMETRIZE TABLE NAMES...BEWARE OF SQL INJECTION ATTACKS
	print (cmd)
	try:
		cursor.execute(cmd, name)
	except :

		print (cursor._last_executed)


	



db = MySQLdb.connect(host = "localhost",
						 user = "Anirudh",
						 passwd = "password",
						 db = "test")

cur = db.cursor()

cur.execute("USE test")
cur.execute("SHOW TABLES")

for (table_name,) in cur:
	print(table_name)


#filename = input("Place ITS Schedule Spreadsheet in the same folder as this program, then type its filename here (e.g Schedule.xlsx): ")
wb = load_workbook("Schedule.xlsx")
scheduleSheet = wb['CarlTech Schedule']
staffSheet = wb['CarlTech Staff List']
workerList = AuxiliaryFunctions.getListOfWorkers(wb)


#initializeDB(db, 'test')

#initializeShiftList(db)
#initializeEmployeeInfoTable(db, workerList)

initializeShiftEmployeeLinker(db, workerList)

#removeWorker(("John", "Doe"))